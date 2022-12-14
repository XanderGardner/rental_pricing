import os
from os.path import exists
import shutil
import sys
import openpyxl as pyxl
from pandas import describe_option
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from threading import Thread
import time
from datetime import datetime
import math
import numpy as np
import bisect

from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager

# INPUT/ CONSTANTS
HEADLESS = False # if running with chrome browser showing (more results when false, but takes longer)

OFFSET = 2 # excel input data is offset by 2: 1 for 0 indexing and 1 for a row of titles
OFFSET_ROWS = 1 # excel input data has 2 extra rows: first is headers

# return all dollar values found given in the dom text
def parse_dollars(dom_text):
    arr_dom_text = dom_text.split()
    dollar_strs = []
    for text in arr_dom_text:
        if text[0] == "$":
            dollar_strs.append(text[1:])
    
    acceptable_chars = {"0", "1", "2", "3", "4", "5", "6", "7", "8", "9", ","}
    dollar_values = []
    for dollar_str in dollar_strs:
        i = 0
        while i < len(dollar_str) and dollar_str[i] in acceptable_chars:
            i += 1
        i -= 1
        # extract numerical value
        multiplier = 1
        value = 0
        while i >= 0:
            if dollar_str[i] == ",":
                i -= 1
                continue
            else:
                value += multiplier * int(dollar_str[i])
                multiplier *= 10
                i -= 1
        dollar_values.append(round(value, 2))
    return dollar_values

# return all dollar values converted to per hour in the given dom text
def parse_timed_dollars(dom_text):
    arr_dom_text = dom_text.split()
    dollar_strs = []
    per_times = []
    for i, text in enumerate(arr_dom_text):
        if text[0] == "$":
            if i+5 <= len(arr_dom_text):
                following_str = ' '.join(arr_dom_text[i:i+5])
            else:
                following_str = ' '.join(arr_dom_text[i:])
            if "hour" in following_str or "hr" in following_str:
                dollar_strs.append(text[1:])
                per_times += [1]
            elif "day" in following_str:
                dollar_strs.append(text[1:])
                per_times += [24]
    
    acceptable_chars = {"0", "1", "2", "3", "4", "5", "6", "7", "8", "9", ","}
    dollar_values = []
    for index, dollar_str in enumerate(dollar_strs):
        i = 0
        while i < len(dollar_str) and dollar_str[i] in acceptable_chars:
            i += 1
        i -= 1
        # extract numerical value
        multiplier = 1
        value = 0
        while i >= 0:
            if dollar_str[i] == ",":
                i -= 1
                continue
            else:
                value += multiplier * int(dollar_str[i])
                multiplier *= 10
                i -= 1
        dollar_values.append(round(value / per_times[index], 2))
    return dollar_values 

# scrape data online for item in row i and set the sourced value
def scrape_sourced_value(row):
    # get variables from excel sheet
    wb = pyxl.load_workbook('temp_delete_me.xlsx', data_only=True)
    ws = wb["General"]

    year = ws[f'B{row}'].value
    description = ws[f'C{row}'].value
    manufacturer = ws[f'E{row}'].value
    model = ws[f'F{row}'].value
    sourced_value = 0.0 if not ws[f'BF{row}'].value else float(ws[f'BF{row}'].value)
    given_value = None if not ws[f'W{row}'].value else float(ws[f'W{row}'].value)
    given_operating_rate = None if not ws[f'H{row}'].value else float(ws[f'H{row}'].value)
    given_standby_rate = None if not ws[f'J{row}'].value else float(ws[f'J{row}'].value)

    wb.close()

    # if sourced value already found, return -1
    if sourced_value:
        return -1
    
    # scrape values online
    sourced_values = []
    def scrape_task(address, sourced_values):
        chrome_options = Options()
        if HEADLESS:
            chrome_options.add_argument("--headless")
        driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()), options=chrome_options)
        driver.get(address)
        
        try:
            WebDriverWait(driver, 10).until(EC.presence_of_element_located(
                (By.TAG_NAME, "body")
            ))
            time.sleep(1.0) # testing shows that an extra second allows all results to load
            
            main_element = driver.find_element(by=By.TAG_NAME, value="body")
            dom_text = main_element.text
            sourced_values += parse_dollars(dom_text)
        finally:
            driver.close()
            driver.quit()
    threads = [None] * 8
    threads[0] = Thread(target=scrape_task, args=(f"https://usedequipmentguide.com/listings?query={description} {manufacturer} {year} used price", sourced_values))
    threads[1] = Thread(target=scrape_task, args=(f"https://www.bing.com/search?q={description} {manufacturer} {year} used price", sourced_values))
    threads[2] = Thread(target=scrape_task, args=(f"https://swisscows.com/web?query={description} {manufacturer} {year} used price", sourced_values))
    threads[3] = Thread(target=scrape_task, args=(f"https://duckduckgo.com/?q={description} {manufacturer} {year} used price&ia=web", sourced_values))
    threads[4] = Thread(target=scrape_task, args=(f"https://search.givewater.com/serp?q={description} {manufacturer} {year} used price", sourced_values))
    threads[5] = Thread(target=scrape_task, args=(f"https://ekoru.org/?q={description} {manufacturer} {year} used price", sourced_values))
    threads[6] = Thread(target=scrape_task, args=(f"https://www.ecosia.org/search?method=index&q={description} {manufacturer} {year} used price", sourced_values))
    threads[7] = Thread(target=scrape_task, args=(f"https://www.google.com/search?q={description} {manufacturer} {year} used price", sourced_values))
    for i in range(8):
        threads[i].start()
    for i in range(8):
        threads[i].join()

    # prepare for output
    wb = pyxl.load_workbook('equipment rates.xlsx')
    ws = wb["General"]

    # corner case: insufficient data found
    sourced_values.sort()
    sourced_values = sourced_values[bisect.bisect_right(sourced_values, 200):] # remove all vals less than 200
    if len(sourced_values) < 3:
        ws[f'BB{row}'] = "Insufficient data available online"
        wb.save('equipment rates.xlsx')
        wb.close()
        return 0

    # save found value to excel

    # mean of the middle 50% of found values
    sourced_value = round(np.median(sourced_values),2)
    
    ws[f'BF{row}'] = str(sourced_value)
    ws[f'BG{row}'] = str(sourced_values)[1:-1]
    ws[f'BH{row}'] = f"https://www.google.com/search?q={description} {manufacturer} {year} used price"
    ws[f'BI{row}'] = str(datetime.now())
    wb.save('equipment rates.xlsx')
    wb.close()

    return 1

# scrape data online for item in row row and set the sourced rental rate
def scrape_sourced_rental_rate(row, location):
    # get variables from excel sheet
    wb = pyxl.load_workbook('temp_delete_me.xlsx', data_only=True)
    ws = wb["General"]

    year = ws[f'B{row}'].value
    description = ws[f'C{row}'].value
    manufacturer = ws[f'E{row}'].value
    model = ws[f'F{row}'].value
    sourced_rental_rate = None if not ws[f'BC{row}'].value else float(ws[f'BC{row}'].value)
    given_value = None if not ws[f'W{row}'].value else float(ws[f'W{row}'].value)
    given_operating_rate = 0.0 if not ws[f'H{row}'].value else float(ws[f'H{row}'].value)
    given_standby_rate = None if not ws[f'J{row}'].value else float(ws[f'J{row}'].value)

    wb.close()

    # if sourced rental rate already found, return -1
    if sourced_rental_rate:
        return -1

    # scrape rental rates online
    sourced_rental_rates = []
    def scrape_task(address, sourced_rental_rates):
        chrome_options = Options()
        if HEADLESS:
            chrome_options.add_argument("--headless")
        driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()), options=chrome_options)
        driver.get(address)
        
        try:
            WebDriverWait(driver, 10).until(EC.presence_of_element_located(
                (By.TAG_NAME, "body")
            ))
            time.sleep(1.0) # testing shows that an extra second allows all results to load
            
            main_element = driver.find_element(by=By.TAG_NAME, value="body")
            dom_text = main_element.text
            sourced_rental_rates += parse_timed_dollars(dom_text)
        finally:
            driver.close()
            driver.quit()
    threads = [None] * 8
    threads[0] = Thread(target=scrape_task, args=(f"https://usedequipmentguide.com/listings?query=how much does it cost to rent a {description} {manufacturer} {year} in {location}", sourced_rental_rates))
    threads[1] = Thread(target=scrape_task, args=(f"https://www.bing.com/search?q=how much does it cost to rent a {description} {manufacturer} {year} in {location}", sourced_rental_rates))
    threads[2] = Thread(target=scrape_task, args=(f"https://swisscows.com/web?query=how much does it cost to rent a {description} {manufacturer} {year} in {location}", sourced_rental_rates))
    threads[3] = Thread(target=scrape_task, args=(f"https://duckduckgo.com/?q=how much does it cost to rent a {description} {manufacturer} {year} in {location}&ia=web", sourced_rental_rates))
    threads[4] = Thread(target=scrape_task, args=(f"https://search.givewater.com/serp?q=how much does it cost to rent a {description} {manufacturer} {year} in {location}", sourced_rental_rates))
    threads[5] = Thread(target=scrape_task, args=(f"https://ekoru.org/?q=how much does it cost to rent a {description} {manufacturer} {year} in {location}", sourced_rental_rates))
    threads[6] = Thread(target=scrape_task, args=(f"https://www.ecosia.org/search?method=index&q=how much does it cost to rent a {description} {manufacturer} {year} in {location}", sourced_rental_rates))
    threads[7] = Thread(target=scrape_task, args=(f"https://www.google.com/search?q=how much does it cost to rent a {description} {manufacturer} {year} in {location}", sourced_rental_rates))
    for i in range(8):
        threads[i].start()
    for i in range(8):
        threads[i].join()

    # prepare for output
    wb = pyxl.load_workbook('equipment rates.xlsx')
    ws = wb["General"]
    sourced_rental_rates.sort()
    sourced_rental_rates = sourced_rental_rates[bisect.bisect_right(sourced_rental_rates, 0):] # remove all 0s and negatives

    # corner case: insufficient data found
    if len(sourced_rental_rates) < 3:
        ws[f'BB{row}'] = "Insufficient data available online"
        wb.save('equipment rates.xlsx')
        wb.close()
        return 0

    # save found rental rate to excel
    # mean of the middle 50% of found values
    sourced_rental_rate = round(np.median(sourced_rental_rates),2)
    
    # find if given_operating_rate fits into given data
    near_average = False
    is_extreme_small = False
    is_extreme_big = False
    if given_operating_rate <= 1.5*sourced_rental_rate and given_operating_rate >= 0.5*sourced_rental_rate:
        near_average = True
    if given_operating_rate < sourced_rental_rates[0]:
        is_extreme_small = True
    elif given_operating_rate > sourced_rental_rates[-1]:
        is_extreme_big = True
    
    # make recommendation based on fit in data
    rec = ""
    if near_average:
        if is_extreme_small:
            rec = "Further research required: possible slight increase in pricing"
        elif is_extreme_big:
            rec = "Further research required: possible slight decrease in pricing"
        else:
            rec = "Data supports pricing"
    else: # not near average
        if is_extreme_small:
            rec = "Consider increase in pricing"
        elif is_extreme_big:
            rec = "Consider decrease in pricing"
        elif given_operating_rate < sourced_rental_rate:
            # not near average but not an extreme
            rec = "Further research required: possible slight increase in pricing"
        else:
            rec = "Further research required: possible slight decrease in pricing"

    ws[f'BB{row}'] = rec
    ws[f'BC{row}'] = str(sourced_rental_rate)
    ws[f'BD{row}'] = str(sourced_rental_rates)[1:-1]
    ws[f'BE{row}'] = f"https://www.google.com/search?q=how much does it cost to rent a {description} {manufacturer} {year} in {location}"
    ws[f'BI{row}'] = str(datetime.now())

    wb.save('equipment rates.xlsx')
    wb.close()

    return 1

def main():
    # output intro screen
    print("")
    print("~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~\n")
    print("           /$$$$$$                                          ")
    print("          /$$__  $$                                         ")
    print("         | $$  \__/  /$$$$$$   /$$$$$$   /$$$$$$   /$$$$$$$ ")
    print("         | $$       /$$__  $$ /$$__  $$ /$$__  $$ /$$_____/ ")
    print("         | $$      | $$$$$$$$| $$  \__/| $$$$$$$$|  $$$$$$  ")
    print("         | $$    $$| $$_____/| $$      | $$_____/ \____  $$ ")
    print("         |  $$$$$$/|  $$$$$$$| $$      |  $$$$$$$ /$$$$$$$/ ")
    print("          \______/  \_______/|__/       \_______/|_______/  \n\n")
    print("~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~\n")
    print("             Rental Pricing Tool developed by Xander Gardner\n\n")
    print("    - Please do NOT use computer while running\n")
    print("    - Expect to wait several hours\n")
    print("    - Refer to rental_pricing_guide.pdf for details\n\n")
    print("Starting in 10 seconds")
    # time.sleep(10)

    # copy excel sheet to new sheet
    original = r'equipment rates.xlsx'
    target = r'temp_delete_me.xlsx'
    if not exists(target):
        shutil.copyfile(original, target)
    
    # get data from 'temp_delete_me.xlsx'
    wb = pyxl.load_workbook('temp_delete_me.xlsx')
    ws = wb["General"]
    n = ws.max_row - OFFSET_ROWS
    location = str(ws[f'BA2'].value)
    wb.close()

    num_to_scrape = n
    # num_to_scrape = 30
    for row in range(OFFSET, num_to_scrape + OFFSET):
        scrape_sourced_value(row)
        scrape_sourced_rental_rate(row, location)

    # remove temp_delete_me
    if os.path.exists(target):
        os.remove(target)

    print("~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~")
    print("~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~")
    print("~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~\n")
    print("COMPLETE!\n\n")
    print("Closing in 30 seconds")
    time.sleep(30)

if __name__ == "__main__":
    main()
